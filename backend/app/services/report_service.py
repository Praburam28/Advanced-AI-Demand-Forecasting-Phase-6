import os
from datetime import datetime

from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models.report import Report
from app.models.forecast import Forecast
from app.models.organization import Organization
from app.schemas.report_schema import ReportCreate
from app.services.activity_service import log_activity
from app.utils.response_handler import success_response


REPORT_DIR = "app/generated_reports"
os.makedirs(REPORT_DIR, exist_ok=True)


def clean_optional_id(value):
    if value is None:
        return None

    try:
        value = int(value)
    except Exception:
        return None

    return value if value > 0 else None


def get_default_organization_id(db: Session):
    organization = db.query(Organization).order_by(Organization.id.asc()).first()

    if not organization:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No organization found. Please create an organization first."
        )

    return organization.id


def create_report(db: Session, data: ReportCreate, current_user):
    forecast = None

    forecast_id = clean_optional_id(data.forecast_id)
    project_id = clean_optional_id(data.project_id)

    if forecast_id:
        forecast = db.query(Forecast).filter(
            Forecast.id == forecast_id,
            Forecast.user_id == current_user.id
        ).first()

        if not forecast:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Forecast not found"
            )

        organization_id = forecast.organization_id
    else:
        organization_id = get_default_organization_id(db)

    report = Report(
        title=data.title,
        report_type=data.report_type,
        forecast_id=forecast_id,
        project_id=project_id,
        file_path=None,
        status="generated",
        created_by=current_user.id,
        organization_id=organization_id,
        report_category="forecast"
    )

    db.add(report)
    db.commit()
    db.refresh(report)

    if report.project_id:
        log_activity(
            db,
            report.project_id,
            current_user.id,
            "Report Generated",
            f"Report '{report.title}' was generated"
        )

    return report


def get_reports(db: Session, current_user, organization_id=None):
    query = db.query(Report).filter(
        Report.created_by == current_user.id
    )

    if organization_id:
        query = query.filter(
            Report.organization_id == organization_id
        )

    return query.order_by(Report.created_at.desc()).all()


def get_report_by_id(db: Session, report_id: int, current_user):
    report = db.query(Report).filter(
        Report.id == report_id,
        Report.created_by == current_user.id
    ).first()

    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report not found"
        )

    return report


def generate_pdf_report(db: Session, report_id: int, current_user):
    report = get_report_by_id(db, report_id, current_user)

    forecast = None
    if report.forecast_id:
        forecast = db.query(Forecast).filter(
            Forecast.id == report.forecast_id,
            Forecast.user_id == current_user.id
        ).first()

    file_path = os.path.join(REPORT_DIR, f"report_{report.id}.pdf")

    doc = SimpleDocTemplate(
        file_path,
        pagesize=A4,
        rightMargin=35,
        leftMargin=35,
        topMargin=35,
        bottomMargin=35
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        textColor=colors.HexColor("#4F46E5"),
        fontSize=24,
        spaceAfter=16,
        alignment=1
    )

    section_style = ParagraphStyle(
        "SectionStyle",
        parent=styles["Heading2"],
        textColor=colors.HexColor("#0F172A"),
        fontSize=15,
        spaceBefore=14,
        spaceAfter=8
    )

    normal_style = ParagraphStyle(
        "NormalStyle",
        parent=styles["BodyText"],
        textColor=colors.HexColor("#334155"),
        fontSize=10,
        leading=15
    )

    story = []

    story.append(Paragraph("AI Demand Forecasting Executive Report", title_style))
    story.append(Paragraph(report.title, section_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d-%m-%Y %I:%M %p')}", normal_style))
    story.append(Spacer(1, 16))

    if forecast:
        kpi_data = [
            ["Metric", "Value"],
            ["Model", forecast.model_name],
            ["Predicted Demand", f"{forecast.predicted_demand} units"],
            ["Revenue Forecast", f"₹{forecast.revenue_forecast}"],
            ["Cost Forecast", f"₹{forecast.cost_forecast}"],
            ["Profit Forecast", f"₹{forecast.profit_forecast}"],
            ["MAPE", f"{forecast.mape}%"],
            ["R2 Score", str(forecast.r2_score)],
        ]
    else:
        kpi_data = [
            ["Metric", "Value"],
            ["Report Type", report.report_type],
            ["Status", report.status],
            ["Forecast Linked", "No forecast linked"],
        ]

    table = Table(kpi_data, colWidths=[2.5 * inch, 3.5 * inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F8FAFC"), colors.HexColor("#EEF2FF")]),
        ("PADDING", (0, 0), (-1, -1), 10),
    ]))

    story.append(Paragraph("Forecast KPI Summary", section_style))
    story.append(table)
    story.append(Spacer(1, 18))

    if forecast:
        story.append(Paragraph("AI Forecast Summary", section_style))
        story.append(Paragraph(forecast.ai_summary or "No AI summary available.", normal_style))

        story.append(Spacer(1, 12))
        story.append(Paragraph("Business Recommendation", section_style))
        story.append(Paragraph(forecast.recommendation or "No recommendation available.", normal_style))

    story.append(Spacer(1, 28))

    footer_data = [
        ["Forecast AI SaaS", "Professional Demand Forecasting Report"],
        ["Prepared For", current_user.email],
    ]

    footer = Table(footer_data, colWidths=[2.5 * inch, 3.5 * inch])
    footer.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ECFEFF")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BAE6FD")),
        ("PADDING", (0, 0), (-1, -1), 9),
    ]))

    story.append(footer)

    doc.build(story)

    report.file_path = file_path
    db.commit()

    return file_path


def generate_excel_report(db: Session, report_id: int, current_user):
    report = get_report_by_id(db, report_id, current_user)

    forecast = None
    if report.forecast_id:
        forecast = db.query(Forecast).filter(
            Forecast.id == report.forecast_id,
            Forecast.user_id == current_user.id
        ).first()

    file_path = os.path.join(REPORT_DIR, f"report_{report.id}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Report"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    sub_fill = PatternFill("solid", fgColor="E0E7FF")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    orange_fill = PatternFill("solid", fgColor="FFEDD5")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=18, bold=True, color="4F46E5")
    bold_font = Font(bold=True, color="0F172A")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws.merge_cells("A1:D1")
    ws["A1"] = "AI Demand Forecasting Executive Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = report.title
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A4"] = "Generated On"
    ws["B4"] = datetime.now().strftime("%d-%m-%Y %I:%M %p")

    ws["A6"] = "Metric"
    ws["B6"] = "Value"

    for cell in ws["6:6"]:
        cell.fill = header_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    if forecast:
        rows = [
            ("Model", forecast.model_name),
            ("Predicted Demand", forecast.predicted_demand),
            ("Revenue Forecast", forecast.revenue_forecast),
            ("Cost Forecast", forecast.cost_forecast),
            ("Profit Forecast", forecast.profit_forecast),
            ("MAE", forecast.mae),
            ("MSE", forecast.mse),
            ("RMSE", forecast.rmse),
            ("MAPE", forecast.mape),
            ("R2 Score", forecast.r2_score),
        ]
    else:
        rows = [
            ("Report Type", report.report_type),
            ("Status", report.status),
            ("Forecast Linked", "No"),
        ]

    start_row = 7
    for index, row in enumerate(rows, start=start_row):
        ws[f"A{index}"] = row[0]
        ws[f"B{index}"] = row[1]

        ws[f"A{index}"].font = bold_font
        ws[f"A{index}"].fill = sub_fill
        ws[f"B{index}"].fill = green_fill if "Profit" in row[0] else orange_fill

        ws[f"A{index}"].border = border
        ws[f"B{index}"].border = border

    if forecast:
        summary_row = start_row + len(rows) + 2

        ws[f"A{summary_row}"] = "AI Summary"
        ws[f"A{summary_row}"].font = white_font
        ws[f"A{summary_row}"].fill = header_fill

        ws.merge_cells(start_row=summary_row + 1, start_column=1, end_row=summary_row + 3, end_column=4)
        ws[f"A{summary_row + 1}"] = forecast.ai_summary or "No AI summary available."
        ws[f"A{summary_row + 1}"].alignment = Alignment(wrap_text=True, vertical="top")

        rec_row = summary_row + 5
        ws[f"A{rec_row}"] = "Business Recommendation"
        ws[f"A{rec_row}"].font = white_font
        ws[f"A{rec_row}"].fill = header_fill

        ws.merge_cells(start_row=rec_row + 1, start_column=1, end_row=rec_row + 3, end_column=4)
        ws[f"A{rec_row + 1}"] = forecast.recommendation or "No recommendation available."
        ws[f"A{rec_row + 1}"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    wb.save(file_path)

    report.file_path = file_path
    db.commit()

    return file_path


def delete_report(db: Session, report_id: int, current_user):
    report = get_report_by_id(db, report_id, current_user)

    project_id = report.project_id
    report_title = report.title

    db.delete(report)
    db.commit()

    if project_id:
        log_activity(
            db,
            project_id,
            current_user.id,
            "Report Deleted",
            f"Report '{report_title}' was deleted"
        )

    return success_response("Report deleted successfully")